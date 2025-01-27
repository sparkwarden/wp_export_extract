#-------------------------------------------------------------
# 
#-------------------------------------------------------------
__all__ = ['build_file_list','list_to_xlsx','xlsx_to_list','LF','File_Node','Message_Writer','get_text_from_file','make_dt_output_filepath','make_std_output_filepath','get_file_hash','put_text_to_file']


#-------------------------------------------------------------
# 
#-------------------------------------------------------------

__prog__ = str(__file__).rstrip('.py')
__author__ = 'Gary D. Smith <https://github.com/sparkwarden>'
__version__ = '3.0'
__date__ = '2025/01/26'

#-------------------------------------------------------------
# 
#-------------------------------------------------------------

"""
Description: File Utility library. Contains message writer, file list builder,
list-to-excel, excel-to-list, and other file-oriented functions.

Tested with Python 3.10
Operating System: iPadOS 17.4
iOS Python apps: Pythonista, a-Shell

Required modules:
 openpyxl - to generate excel report file.
 
Change Summary:
	3.0 simplified 'Message_Writer'.  defer_msg replaced by buffer_msg.  Instance creator requires output logfile path name.  Messages written to file or buffered  explicitly.  In other words, you can buffer messages then write them.  Msgs no longer auto-buffered and written when buffer limit reached.
	2.0 deprecated 'Buffer_Writer' class and related functions 'make_dt_output_fd' and 'make_std_output_fd'.  Use 'Message_Writer' instead.
 
"""

#-------------------------------------------------------------
# 
#-------------------------------------------------------------

import pathlib
import datetime
import io
import sys
import openpyxl
from operator import attrgetter
import mimetypes
import time
import hashlib


LF = '\n'

#-------------------------------------------------------------
# 
#-------------------------------------------------------------

def get_file_hash(file_path, chunksize=4096):
	hash_md5 = hashlib.md5()
	with open(file_path, "rb") as f:
		for chunk in iter(lambda: f.read(chunksize), b""): 
			hash_md5.update(chunk)
	return hash_md5.hexdigest()
	
#-------------------------------------------------------------
# 
#-------------------------------------------------------------

class File_Node:
	"""
	Models file attributes and behavior.
	"""
	
	dir_set = set()
	filetype_set = set()
	
	filenode_list = []
	sorted_filenode_list = []
	
	#-------------------------------------------------------------
	#   
	#-------------------------------------------------------------  
	
	def __new__(cls, path):
		return super().__new__(cls)
		
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	def __init__(self,path):
		
		_cls = File_Node
		
		self.path = path
		p = pathlib.Path(path)
		self.parents = list(p.parents)
		_stat = p.stat()
		self.drive = str(p.drive) 	# file drive, used mainly on windows
		self.ext = str(p.suffix)		# file extension
		self.parentdir = str(p.parent) # parent directory
		self.filename = str(p.name)		# filename, stem + suffix
		
		_dt_fmt = '{:%Y-%m-%d %H:%M:%S}'
		_dt_fmt = '%a %b %d %H:%M:%S %Y'
		
		_created = time.ctime(_stat.st_ctime)
		_modified = time.ctime(_stat.st_mtime)
		_accessed = time.ctime(_stat.st_atime)
	
		self.dt_created = datetime.datetime.strptime(_created, _dt_fmt)
		self.dt_modified = datetime.datetime.strptime(_modified, _dt_fmt)
		self.dt_accessed = datetime.datetime.strptime(_accessed, _dt_fmt)
		
		self.dt_str_created = _dt_fmt.format(_created)
		self.dt_str_modified =_dt_fmt.format(_modified)
		self.dt_str_accessed = _dt_fmt.format(_accessed)
		
		self.filesize = _stat.st_size
		
		self.sortkey = self.path
		
		try:
			_file_type = mimetypes.types_map[self.ext]
		except (KeyError, Exception):
			_file_type = ''
		finally:
			self.filetype = _file_type

		self.is_symlink = p.is_symlink()
		self.is_hardlink = _stat.st_nlink > 1
		
		self.is_in_trash = False
		
		_cls.dir_set.add(self.parentdir)
		_cls.filetype_set.add(self.filetype)
			
		self.text_content = ''
		
		if self not in _cls.filenode_list:
			_cls.filenode_list.append(self)
			_cls.sorted_filenode_list.append(self)
			
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	def set_text_content(self):
		p = pathlib.Path(self.path)
		if self.filetype.startswith('text'):
			self.text_content = p.read_text(encoding='utf-8')
		else:
			self.text_content = ''
			
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	def as_dict(self):
		return self.__dict__
	
	#------------------------------------------------------------------
	#
	#------------------------------------------------------------------
	
	def as_str(self):
		_msg = f'\n\n<{__class__.__name__}> '
		for k,v in self.as_dict().items():
			_msg += f'\n{k}: {v} '
		return _msg
	
	#------------------------------------------------------------------
	#
	#------------------------------------------------------------------
	
	def __repr__(self) -> str:
		return self.as_str()
		
	#------------------------------------------------------------------
	#
	#------------------------------------------------------------------
	
	
	@classmethod
	def sort_nodes(cls,sort_reversed=False):
		_filenodes = cls.filenode_list
		cls.sorted_filenode_list = sorted(_filenodes,key=attrgetter('sortkey'),reverse=sort_reversed)
		
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	@classmethod
	def get_filenode_from_path(cls,path):
		retnode = None
		
		_ret_node_list = [fn for fn in cls.filenode_list if fn.path == path]
		if len(_ret_node_list) == 1:
			retnode = _ret_node_list[0]
			
		return retnode
		
#-------------------------------------------------------------
# 
#-------------------------------------------------------------
		
File_Node = File_Node
			
#-------------------------------------------------------------
# 
#-------------------------------------------------------------

class Message_Writer:
	"""
	Output Messages to List (buffer) then to file.
	"""
	
	msg_node_list = []
	
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	def __init__(self, name:str, file_path:str,echo_to_console=True,file_mode='a'):
		self.name = name
		self.msg_buffer = []
		self.echo_to_console = echo_to_console
		self.file_mode = file_mode
		self.file_path = file_path
		if self not in Message_Writer.msg_node_list:
			Message_Writer.msg_node_list.append(self)
			
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	def write_msg(self,s):
		msg_list = [f'{LF}{s}']
		self.write_msg_from_list(msg_list)
		if self.echo_to_console:
			print(s)
		
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
		
	def buffer_msg(self,s):
		self.msg_buffer.append(s)
		if self.echo_to_console:
			print(s)
			
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
			
	def write_messages_from_buffer(self):
		if len(self.msg_buffer) > 0:
			msg_list = self.msg_buffer
			self.write_msg_from_list(msg_list)
			self.msg_buffer.clear()
		
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
			
	def close(self):
		self.write_messages_from_buffer()
				
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	def write_msg_from_list(self,msg_list:list,skip_type_err=True):
	
		with open(self.file_path, mode=self.file_mode, encoding='utf-8') as tf:
			for txt in msg_list:
				try:
					tf.writelines(txt)
				except TypeError as ex:
					if skip_type_err:
						pass
					else:
						raise TypeError(ex)
						
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	@classmethod
	def shutdown(cls):
		for nd in cls.msg_node_list:
			nd.close()
				
#-------------------------------------------------------------
#
#-------------------------------------------------------------

class Message_Writer_Orig:
	"""
	Output Messages to File.
	"""
	
	node_list = []
	
	PRN_SCREEN_AND_FILE = 1
	PRN_SCREEN_ONLY = 2
	PRN_FILE_ONLY = 3
	PRN_FLUSH_LOG_THRESHOLD = 100
	
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	def __new__(cls, name, prefix, prn_flag=1):
		return super().__new__(cls)
		
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	def __init__(self, name, prefix, prn_flag=1):
		
		_cls = Message_Writer
		
		self.is_active = False
		self.name = name
		self.prn_flag = prn_flag
		self.msg_cnt = 0
			
		self.msgfilepath = _cls.get_output_path_with_dt(prefix)
			
		self.msgbuf = io.StringIO()
		self.deferred_message_list = []
			
		if self not in _cls.node_list:
			_cls.node_list.append(self)
			msg = f'{LF}writer [{self.name}] started. . .\n'
			with open(self.msgfilepath,'w',encoding='utf-8') as f:
				f.write('')
				self.write_msg(msg)
		else:
			raise Exception ('message writer already established')
			
		self.is_active = True
		
		
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	@staticmethod
	def get_output_path_with_dt(prefix='msgr', ext='.log',\
		_timefmt='%Y%m%d_%H%m%S%f'):
		_path = ''.join([prefix, '_', datetime.datetime.now().strftime(_timefmt), ext])
		return _path
	
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	@classmethod
	def get_writer_by_name(cls, name):
		"""
		return writer node by name.
		"""
		
		_node_list = [nd for nd in cls.node_list if nd.name == name]
		retnode = None
		if len(_node_list) > 0:
			retnode = _node_list[0]
		return retnode
	
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	def defer_msg(self, msg):
		"""
		Write message to deferred list.
		"""
		self.deferred_message_list.append(msg)
	
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	def write_deferred_msgs(self):
		"""
		Write deferred messages.  
		"""
	
		for msg in self.deferred_message_list:
			self.write_msg(msg)
		self.deferred_message_list.clear()
	
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	def write_msg(self, *args, skiplines=1):
		"""
		write msg to buffer.
		"""
		_arg_str = f'{LF*skiplines}'
		for arg in args:
			_arg_str += str(arg)
		
		_cls = Message_Writer
		
		if self.msg_cnt >= _cls.PRN_FLUSH_LOG_THRESHOLD:
			self.flushbuf()
		
		if self.prn_flag == _cls.PRN_SCREEN_AND_FILE:
			self.msgbuf.write(_arg_str)
			sys.stdout.write(_arg_str)
		elif self.prn_flag == _cls.PRN_SCREEN_ONLY:
			sys.stdout.write(_arg_str)
		else:
			self.msgbuf.write(_arg_str)
			
		self.msg_cnt+=1
		
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
			
	def flushbuf(self):
		"""
		write the msgbuf to file.
		"""
		with open(self.msgfilepath,'a',encoding='utf-8') as f:
			f.write(self.msgbuf.getvalue())
		self.msg_cnt = 0
		
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------

	def close_writer(self):
		"""
		Write accumulated messages from msgbuf to output file.  
		Note: deferred messages must be written manually using
		'write_deferred_msgs'.
		"""
		self.write_msg(f'writer [{self.name}] closed. . .')
		self.flushbuf()
		self.msgbuf.close()
		self.is_active = False
	
	#-------------------------------------------------------------
	# 
	#------------------------------------------------------------
		
	@classmethod
	def setup(cls):
		cls.node_list.clear()
		
	#-------------------------------------------------------------
	# 
	#------------------------------------------------------------
		
	@classmethod
	def shutdown(cls):
		for lg in cls.node_list:
			if lg.is_active:
				lg.close_writer()
	
#-------------------------------------------------------------
# 
#-------------------------------------------------------------

def get_text_from_file(path):
	text_rows = []
	with open(path, 'r', encoding='utf-8') as tf:
		text_rows = tf.readlines()
	return text_rows

#-------------------------------------------------------------
# 
#-------------------------------------------------------------

def put_text_to_file(text_list, text_path):
	
	with open(text_path, 'w', encoding='utf-8') as tf:
		for txt in text_list:
			tf.writelines(txt)

#-------------------------------------------------------------
# 
#-------------------------------------------------------------


def list_to_xlsx(xls_list,xls_path,xls_sheet=None):
	"""
	Write list to excel .xlsx file
	"""
	
	wb = openpyxl.Workbook(write_only=True)
	if xls_sheet:
		ws = wb.create_sheet(xls_sheet)
	else:
		ws = wb.create_sheet()
	
	for row in xls_list:
		ws.append(row)

	wb.save(xls_path)

	
#-------------------------------------------------------------
# 
#-------------------------------------------------------------
	
def xlsx_to_list(xlsx_path):
	"""
	Return a nested list from an excel .xlsx file.
	Structure: [[rows[cells]]]
	"""
	wb = openpyxl.load_workbook(filename=xlsx_path, read_only=True)
	ws = wb.active
	xlsx_list = []
	for row in ws.rows:
		row_list = []
		for cell in row:
			row_list.append(str(cell.value))
			
		if len(row_list) > 0:
			xlsx_list.append(row_list)
	
	wb.close()
	return xlsx_list
		
#-------------------------------------------------------------
# 
#-------------------------------------------------------------

def build_file_list(startdir:str=None, ptrnstr='*.*') -> list:
	"""
	Return a list of files recursively, starting with
	[startdir] directory, matching [ptrnstr] search pattern.
	"""
	_file_list = []
	
	if startdir is None:
		_startdir = str(pathlib.Path().cwd())
	else:
		_startdir = startdir
	
	p = pathlib.Path(_startdir)
	
	glob_list = p.rglob(ptrnstr)
	
	for path in glob_list:
		_path = str(path)
		p = pathlib.Path(_path)
		
		if p.is_file(): 
			_file_list.append(_path)
	
	return _file_list

#-------------------------------------------------------------
# 
#-------------------------------------------------------------


def make_dt_output_filepath(prefix='logfile', ext='.log',\
	timefmt='%Y%m%d_%H%m%S%f'):
	"""
	Return str output filepath using name prefix, datetime format
	and file extension. 
	"""
	out_path = prefix + '_' + datetime.datetime.now().strftime(timefmt) + ext
	return out_path


#-------------------------------------------------------------
# 
#-------------------------------------------------------------

def make_std_output_filepath(prefix='logfile', ext='.log'):
	"""
	Return str output filepath using name prefix and file extension. 
	"""
	return prefix + ext
	
#-------------------------------------------------------------
# 
#-------------------------------------------------------------

def cut_list_segment(in_list:list,start_str:str,end_str:str):
	out_list = []
	out_on = False
	
	for txt in in_list:
		if start_str in txt:
			out_on = True
		if end_str in txt:
			out_on = False
		if out_on:
			out_list.append(txt)
	
	return out_list

#-------------------------------------------------------------
# 
#-------------------------------------------------------------

def merge_text_from_files(start_dir:str):
	text_list = []
	
	text_file_list = build_file_list(start_dir,'*.txt')
	
	for path in text_file_list:
		txt = get_text_from_file(path)
		text_list += txt
		
	return text_list
	
#-------------------------------------------------------------
# 
#-------------------------------------------------------------

g_msgwr = None

def write_msg(s):
	g_msgwr.write_msg(s)

#-------------------------------------------------------------
# 
#-------------------------------------------------------------

def main():
	
	#-------------------------------------------------------------
	# 
	#------------------------------------------------------------
	
	curdir = str(pathlib.Path().cwd())
	
	write_msg(f'curdir: {curdir}')
	
	file_list = build_file_list(curdir, '*.py')
	
	for filepath in file_list:
		fn = File_Node(filepath)
		fn.sortkey = str(fn.filename).lower()
	
	write_msg(f'{len(file_list)} [.py] files in {curdir} dir tree. {LF}')
	
	File_Node.sort_nodes()
	
	for fn in File_Node.sorted_filenode_list:
		_matches_found = False
		_match_strs = ['__new__']
		_source_str = ''
		_parentdir = str(fn.parentdir).lower()
		_is_save_file = (('archive' in _parentdir) or ('save' in _parentdir))
		if _is_save_file:
			continue
		
		if str(fn.filetype).startswith('text'):
			fn.set_text_content()
			_source_str = str(fn.text_content).lower()
			for s in _match_strs:
				if s in _source_str:
					_matches_found = True
				
		if _matches_found:
			write_msg(f'path: {fn.path} match_found: {_matches_found}')
	
	#-------------------------------------------------------------
	# 
	#------------------------------------------------------------
	
	
#-------------------------------------------------------------
# 
#------------------------------------------------------------

if __name__ == "__main__":
	
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	msg_log_file = make_dt_output_filepath(prefix=__prog__, ext='.log')
	
	g_msgwr = Message_Writer(name='root',file_path=msg_log_file)
	
	write_msg(f'program {__file__} started. {LF}')
	
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	main()
	
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	#-------------------------------------------------------------
	# output a simple xlsx spreadsheet file.
	#-------------------------------------------------------------
	
	xlsx_test_list = [['col1','col2','col3']]

	xlsx_test_path = make_dt_output_filepath(__prog__,'.xlsx')
	
	for r in range(10):
		row = []
		for c in range(3):
			row.append(str(c))
		xlsx_test_list.append(row)
	
	list_to_xlsx(xlsx_test_list,xlsx_test_path)
	
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------
	
	write_msg(LF)
	write_msg(f'program {__file__} completed. {LF}')
	
	#-------------------------------------------------------------
	# 
	#-------------------------------------------------------------

	Message_Writer.shutdown()